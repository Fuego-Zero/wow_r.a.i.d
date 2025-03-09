from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import random

import copy

# 常量定义：
TANK_NEED = {'防骑': 2, '血DK': 1}
HEALER_MAX = 4
HEALER_MIN = 4
HEALER_MUST_HAVE = {'奶骑': 2, '戒律牧': 1}

DPS_LIMIT = {
    '增强萨': (1, 1), '盗贼': (1, 1),
    '惩戒骑': (1, 1), '暗牧': (1, 1), '痛苦术': (1, 2), '恶魔术': (1, 1), '鸟德': (1, 2),
    '猎人': (2, 3), '邪DK': (1, 2), '冰DK': (0, 1),
    '猫德': (0, 1), '电萨': (1, 1), '法师': (2, 3),
    '战士': (1, 1),
}

HEALER_LIMIT = {
    '奶骑': 2,  # 奶骑人数上限（不超过2）
    '戒律牧': 2,  # 戒律牧人数上限（不超过2）
    '奶德': 1,
    '奶萨': 1
}

ROLES = {
    "防骑": "坦克", "血DK": "坦克",
    "战士": "近战输出", "增强萨": "近战输出", "惩戒骑": "近战输出",
    "盗贼": "近战输出", "猫德": "近战输出", "邪DK": "近战输出", "冰DK": "近战输出",
    "猎人": "远程输出", "电萨": "远程输出", "暗牧": "远程输出",
    "痛苦术": "远程输出", "恶魔术": "远程输出", "法师": "远程输出", "鸟德": "远程输出",
    "奶萨": "治疗", "奶骑": "治疗", "戒律牧": "治疗", "奶德": "治疗"
}

# 稀缺职业
SCARCE_ROLES = []

# 颜色映射定义（职业名->颜色）
CLASS_COLORS = {
    '战士': 'CD853F',  # 棕色
    '防骑': 'FFC0CB', '惩戒骑': 'FFC0CB', '奶骑': 'FFC0CB',  # 粉色
    '增强萨': '1E90FF', '电萨': '1E90FF', '奶萨': '1E90FF',  # 蓝色
    '猎人': '98FB98',  # 绿色
    '暗牧': '000000', '戒律牧': '000000',  # 黑色
    '痛苦术': 'DA70D6', '恶魔术': 'DA70D6',  # 紫色
    '盗贼': 'FFFF00',  # 黄色
    '法师': '00BFFF',  # 浅蓝色
    '猫德': 'FFA500', '鸟德': 'FFA500', '奶德': 'FFA500',  # 橙色
    '邪DK': 'FF4500', '冰DK': 'FF4500', '血DK': 'FF4500'  # 红色
}


# 数据结构
class Player:
    def __init__(self, player_name, characters, available_times):
        self.player_name = player_name
        self.characters = characters  # [{'name':角色名, 'class':职业名},...]
        self.available_times = available_times


# 在第一次排团开始前构建本次CD周期总角色池
def build_cd_role_pool(players):
    role_pool = dict()
    for player in players:
        role_pool[player.player_name] = copy.deepcopy(player.characters)
    return role_pool


def read_excel(file):
    df = pd.read_excel(file)
    players = []
    for _, row in df.iterrows():
        if pd.isna(row['玩家名称']) or pd.isna(row['角色名和职业']) or pd.isna(row['可以玩的时间']):
            break
        player_name = str(row['玩家名称']).strip()
        char_data = row['角色名和职业'].split('\n')
        characters = []
        for char in char_data:
            if "/" not in char:
                continue  # 跳过格式不标准的角色信息
            name, job = char.strip().split('/')
            characters.append({'name': name.strip(), 'class': job.strip()})
        available_times_raw = str(row['可以玩的时间']).strip()
        available_times = parse_available_times(available_times_raw)
        players.append(Player(player_name, characters, available_times))
    return players


def parse_available_times(available_str):
    slots = []
    time_map = {
        '周四': ['周四-19:30', '周四-20:30'],
        '周五': ['周五-19:30', '周五-20:30'],
        '周六': ['周六-19:30', '周六-20:30'],
        '周日': ['周日-19:30', '周日-20:30'],
        '周一': ['周一-19:30', '周一-20:30']
    }
    parts = available_str.split('、')
    for part in parts:
        part = part.strip()
        if '第二车' in part:
            day = part.replace('第二车', '').strip()
            if day in time_map:
                slots.append(time_map[day][1])  # 第二车，只能参加20:30这一趟
        else:
            if part in time_map:
                slots.extend(time_map[part])  # 当天两个档（第一车和第二车）都可以参加
    return slots


# 获取该角色所有职业(双修支持)
def get_classes(char_class):
    return char_class.strip().split("+")


# 检查团本配置合法性
def validate_roster(roster):
    tank = defaultdict(int)
    heal = defaultdict(int)
    dps = defaultdict(int)
    for player, char, cls in roster:
        role = ROLES[cls]
        if role == '坦克':
            tank[cls] += 1
        elif role == '治疗':
            heal[cls] += 1
        else:
            dps[cls] += 1

    # 检查坦克是否满足
    for cls, num in TANK_NEED.items():
        if tank[cls] != num:
            return False

    # 检查治疗数量和必须职业
    if not (HEALER_MIN <= sum(heal.values()) <= HEALER_MAX):
        return False
    for cls, num in HEALER_MUST_HAVE.items():
        if heal[cls] < num:
            return False

    # 检查输出职业限制条件
    for cls, (min_n, max_n) in DPS_LIMIT.items():
        if dps[cls] < min_n or dps[cls] > max_n:
            return False

    # 检查远程近战尽可能均衡
    melee_count = sum([v for k, v in dps.items() if ROLES[k] == "近战输出"])
    ranged_count = sum([v for k, v in dps.items() if ROLES[k] == "远程输出"])
    if abs(melee_count - ranged_count) > 2:  # 允许小幅偏差
        return False

    return True


# 排团函数
def create_group(players, cd_role_pool, time_slot, all_time_slots, require_players):
    available_pool = {}
    player_dict = {p.player_name: p for p in players}
    for pname, chars in cd_role_pool.items():
        player = player_dict[pname]
        if time_slot in player.available_times and chars:
            available_pool[pname] = copy.deepcopy(chars)

    selected_roles = []

    def pick_role(pname, cls_needed):
        if pname not in available_pool:
            return False
        for char in available_pool[pname]:
            if cls_needed in get_classes(char['class']):
                selected_roles.append((pname, char['name'], cls_needed))
                del available_pool[pname]
                return True
        return False

    def pick_with_dualspec_priority(cls_needed):
        items = list(available_pool.items())
        random.shuffle(items)
        for pname, chars in items:
            char_list = chars[:]
            random.shuffle(char_list)
            for char in char_list:
                classes = get_classes(char['class'])
                roles_set = set(ROLES[c] for c in classes)
                if cls_needed in classes and len(roles_set) == 1:
                    return pick_role(pname, cls_needed)
        for pname in random.sample(list(available_pool), len(available_pool)):
            if pick_role(pname, cls_needed):
                return True
        return False

    def get_remaining_counts():
        remaining_count = {}
        for player in players:
            remaining_count[player.player_name] = len([slot for slot in player.available_times if slot != time_slot])
        return remaining_count

    remaining_counts = get_remaining_counts()

    priority_classes = list(TANK_NEED.keys()) + \
                       list(HEALER_MUST_HAVE.keys()) + ['奶德', '奶萨'] + \
                       SCARCE_ROLES + \
                       list(DPS_LIMIT.keys())

    dps_selected = defaultdict(int)
    healer_counts = defaultdict(int)  # 职业计数器（治疗）
    total_healers = 0  # 治疗总计数器

    # 治疗职业优先明确顺序处理，且玩家随机排序（奶骑2 > 戒律牧1 > 其他）
    treatment_priority_order = [
        ('奶骑', HEALER_MUST_HAVE['奶骑']),
        ('戒律牧', HEALER_MUST_HAVE['戒律牧']),
        ('奶德', HEALER_LIMIT['奶德']),
        ('奶萨', HEALER_LIMIT['奶萨'])
    ]

    dk_classes = ['邪DK', '冰DK']

    # 已选中的玩家名单，防止重复选择
    used_players = set()

    # 遍历治疗职业，优先池随机顺序分配治疗职业
    for heal_cls, required_number in treatment_priority_order:
        current_cls_count = len([r for r in selected_roles if r[2] == heal_cls])
        while current_cls_count < required_number and total_healers < HEALER_MAX:
            found = False
            # 每次都重新随机打乱玩家顺序，确保每次排团都有随机体验
            shuffled_require_players = [p for p in require_players if p in available_pool and p not in used_players]
            random.shuffle(shuffled_require_players)  # ⭐随机玩家顺序⭐
            for pname in shuffled_require_players:
                if pick_role(pname, heal_cls):
                    healer_counts[heal_cls] += 1
                    total_healers += 1
                    used_players.add(pname)
                    found = True
                    break
            if not found:
                # 若遍历全部玩家都没找到匹配的治疗，则终止当前治疗职业选择
                break
            current_cls_count += 1

    # 后续的坦克和DPS选择，保持原有剔除已选玩家逻辑
    for pname in require_players:
        if pname not in available_pool or pname in used_players:
            continue
        picked = False
        shuffled_priority_classes = priority_classes[:]
        random.shuffle(shuffled_priority_classes)
        for cls in shuffled_priority_classes:
            role = ROLES.get(cls)

            if role == '治疗':
                continue  # 治疗前面已经明确地选完了，不再重复

            elif role == '坦克':
                current_tank_count = len([r for r in selected_roles if r[2] == cls])
                if current_tank_count >= TANK_NEED.get(cls, 0):
                    continue
                if pick_role(pname, cls):
                    picked = True
                    used_players.add(pname)
                    break

            else:  # DPS选择严格检查数量 + 输出DK总数≤2
                current_dps_count = len([r for r in selected_roles if r[2] == cls])
                max_dps_limit = DPS_LIMIT.get(cls, (0, 99))[1]
                if current_dps_count >= max_dps_limit:
                    continue
                if cls in dk_classes and sum(dps_selected[dk] for dk in dk_classes) >= 2:
                    continue
                if pick_role(pname, cls):
                    dps_selected[cls] += 1
                    picked = True
                    used_players.add(pname)
                    break

        if picked and pname in available_pool:
            del available_pool[pname]

    # 坦克选择
    for tank_cls, num in TANK_NEED.items():
        count = len([role for role in selected_roles if role[2] == tank_cls])
        while count < num:
            items = list(available_pool.items())
            random.shuffle(items)
            found = False
            for pname, _ in items:
                if pick_with_dualspec_priority(tank_cls):
                    found = True
                    break
            if not found:
                selected_roles.append(("", "", tank_cls))
            count += 1

    # 治疗选择完美优化版（易读+强制顺序明确版）

    items_healers = list(available_pool.items())
    random.shuffle(items_healers)

    total_healers = len([r for r in selected_roles if ROLES.get(r[2]) == '治疗'])

    # 第一步：严格按照先奶骑2人，再戒律牧1人的顺序，明确选择必须治疗
    ordered_healers_must_have = [('奶骑', HEALER_MUST_HAVE['奶骑']), ('戒律牧', HEALER_MUST_HAVE['戒律牧'])]
    for heal_cls, required_num in ordered_healers_must_have:
        current_count = len([r for r in selected_roles if r[2] == heal_cls])
        while current_count < required_num:
            found = False
            for pname, _ in items_healers:
                current_cls_count = len([r for r in selected_roles if r[2] == heal_cls])
                if current_cls_count >= HEALER_LIMIT.get(heal_cls, HEALER_MAX):
                    continue
                if pick_with_dualspec_priority(heal_cls):
                    total_healers += 1
                    found = True
                    break
            if not found:
                selected_roles.append(("", "", heal_cls))
                total_healers += 1
            current_count += 1

    # 第二步：补充额外治疗直到满4人
    additional_healers = ['戒律牧', '奶萨', '奶德']
    while total_healers < HEALER_MIN:
        random.shuffle(additional_healers)
        found = False
        for heal_cls in additional_healers:
            current_heal_cls_count = len([r for r in selected_roles if r[2] == heal_cls])
            if current_heal_cls_count >= HEALER_LIMIT.get(heal_cls, HEALER_MAX):
                continue
            if pick_with_dualspec_priority(heal_cls):
                total_healers += 1
                found = True
                break
        if not found:
            selected_roles.append(("", "", "治疗"))
            total_healers += 1

    # 第三步：如果意外超过HEALER_MAX(4)，用空缺替代多余治疗
    if total_healers > HEALER_MAX:
        actual_heal_count = 0
        fixed_roles = []
        for role in selected_roles:
            if ROLES.get(role[2]) == '治疗':
                if actual_heal_count < HEALER_MAX:
                    fixed_roles.append(role)
                    actual_heal_count += 1
                else:
                    fixed_roles.append(("", "", "治疗"))  # 超额治疗改为空位
            else:
                fixed_roles.append(role)
        selected_roles = fixed_roles

    # DK特殊处理 (修复版本，严格检查 DPS_LIMIT并确保输出DK不超过2人)
    dk_selected = {'邪DK': 0, '冰DK': 0}
    dk_classes = ['邪DK', '冰DK']

    # 先挑选邪DK玩家
    xiedk_candidates = []
    for pname in available_pool:
        if any('邪DK' in get_classes(c['class']) for c in available_pool[pname]):
            xiedk_candidates.append((remaining_counts[pname], pname))
    random.shuffle(xiedk_candidates)

    max_xiedk = DPS_LIMIT['邪DK'][1]  # 获取邪DK的上限人数

    for _, pname in xiedk_candidates:
        # 增添额外总数检查条件
        if dps_selected['邪DK'] >= max_xiedk or sum(dps_selected[dk] for dk in dk_classes) >= 2:
            break
        if pick_role(pname, '邪DK'):
            dps_selected['邪DK'] += 1
            dk_selected['邪DK'] += 1

    # 如果未选到足够的邪DK，则采用占位补充，但严格检查当前DK总数，确保不超过2
    while dps_selected['邪DK'] < DPS_LIMIT['邪DK'][0]:
        # 👇再增加一次总人数检查！这是第二个关键防护措施！
        if sum(dps_selected[dk] for dk in dk_classes) >= 2:
            break
        selected_roles.append(("", "", "邪DK"))
        dps_selected['邪DK'] += 1

    # DPS基础需求
    dps_min_requirements = {cls: limit[0] for cls, limit in DPS_LIMIT.items() if limit[0] > 0}
    for dps_cls, min_num in dps_min_requirements.items():
        while dps_selected[dps_cls] < min_num:
            found = False
            player_list = list(available_pool.keys())
            random.shuffle(player_list)  # 随机打乱玩家顺序
            for pname in player_list:
                if pick_role(pname, dps_cls):
                    dps_selected[dps_cls] += 1
                    found = True
                    break
            if not found:
                selected_roles.append(("", "", dps_cls))
                dps_selected[dps_cls] += 1

    vacancy_count = 25 - len(selected_roles)
    selected_players = set(pname for pname, _, _ in selected_roles if pname)
    dk_classes = ['邪DK', '冰DK']
    available_dps_classes = [cls for cls, role in ROLES.items() if
                             role in ['近战输出', '远程输出'] and cls not in dk_classes]

    available_dps_characters = []
    for pname, chars in available_pool.items():
        if pname in selected_players:
            continue
        for char in chars:
            char_classes = get_classes(char['class'])
            dps_classes = [cls for cls in char_classes if cls in available_dps_classes]
            if dps_classes:
                available_dps_characters.append((pname, char['name'], dps_classes))

    random.shuffle(available_dps_characters)

    for _ in range(vacancy_count):
        success = False
        while available_dps_characters:
            pname, char_name, dps_classes = available_dps_characters.pop()
            random.shuffle(dps_classes)
            picked = False
            for dps_cls in dps_classes:
                if pick_role(pname, dps_cls):
                    selected_players.add(pname)
                    if pname in cd_role_pool:
                        cd_role_pool[pname] = [c for c in cd_role_pool[pname] if c['name'] != char_name]
                    available_pool.pop(pname, None)
                    success = True
                    picked = True
                    break
            if picked:
                break

        if not success:
            random_cls = random.choice(available_dps_classes)
            selected_roles.append(("", "", random_cls))

    while len(selected_roles) < 25:
        selected_roles.append(("", "", ""))

    for pname, char_name, _ in selected_roles:
        if pname and char_name:
            cd_role_pool[pname] = [c for c in cd_role_pool[pname] if c['name'] != char_name]

    return selected_roles


def export_schedule_to_excel(schedule_dict, players, cd_role_pool, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "团本安排汇总"

    player_lookup = {p.player_name: p for p in players}

    # 准备玩家时间简称显示规则
    def simplify_player_times(player):
        times = player.available_times
        day_all_slots = {'周四': ['周四-19:30', '周四-20:30'],
                         '周五': ['周五-19:30', '周五-20:30'],
                         '周六': ['周六-19:30', '周六-20:30'],
                         '周日': ['周日-19:30', '周日-20:30'],
                         '周一': ['周一-19:30', '周一-20:30']}
        simplified_days = []
        for day, slots in day_all_slots.items():
            if set(slots).issubset(set(times)):
                simplified_days.append(day)
            elif slots[1] in times and slots[0] not in times:
                simplified_days.append(f"{day}第二车")
            elif slots[0] in times and slots[1] not in times:
                simplified_days.append(f"{day}第一车")
        return "，".join(simplified_days)

    player_time_disp = {p.player_name: simplify_player_times(p) for p in players}

    time_slots_order = ["周四-19:30", "周四-20:30", "周五-19:30", "周五-20:30",
                        "周六-19:30", "周六-20:30", "周日-19:30", "周日-20:30",
                        "周一-19:30", "周一-20:30"]

    result = []
    row_cursor = 1

    for slot in time_slots_order:
        row_val = {
            "time": slot,
            "players": []
        }
        ws.cell(row=row_cursor, column=1, value=slot).font = ws.cell(row=row_cursor, column=1).font.copy(bold=True)
        roster = schedule_dict[slot]

        arranged = [r for role in ['坦克', '近战输出', '远程输出', '治疗']
                    for r in roster if ROLES.get(r[2]) == role]

        idx = 0
        for col in range(1, 6):
            for row in range(row_cursor + 1, row_cursor + 6):
                if idx >= len(arranged):
                    continue
                pname, cname, cls = arranged[idx]
                cell_text = f"{pname}-{cname}-{cls}" if pname else f"空缺\n\n{cls}"
                ws.cell(row=row, column=col, value=cell_text).fill = PatternFill(
                    "solid", fgColor=CLASS_COLORS.get(cls, 'FFFFFF') if cls else "FFFFFF"
                )
                idx += 1
                row_val['players'].append({
                    "pname": pname,
                    "cname": cname,
                    "cls": cls
                })
        row_cursor += 7
        result.append(row_val)

        # 未安排玩家
        ws.cell(row=row_cursor, column=1, value=f"{slot} 未安排玩家及可用角色").font = ws.cell(row=row_cursor, column=1).font.copy(bold=True, color="FF0000")
        row_cursor += 1
        assigned_pnames = {pname for pname, _, _ in roster if pname}
        has_unassigned = False

        for player in players:
            if slot not in player.available_times or player.player_name in assigned_pnames:
                continue
            available_chars = [char for char in cd_role_pool.get(player.player_name, [])
                               if slot in player.available_times]
            if available_chars:
                has_unassigned = True
                pname_display = f"{player.player_name}（{player_time_disp[player.player_name]}）"
                ws.cell(row=row_cursor, column=1, value=pname_display)
                col_cursor = 2
                for char in available_chars:
                    main_cls = get_classes(char['class'])[0]
                    ws.cell(row=row_cursor, column=col_cursor, value=f"{char['name']}-{char['class']}").fill =\
                        PatternFill("solid", fgColor=CLASS_COLORS.get(main_cls,'FFFFFF'))
                    col_cursor += 1
                row_cursor += 1
        if not has_unassigned:
            ws.cell(row=row_cursor, column=1, value="✅无")
            row_cursor += 1
        row_cursor += 1

        # 已安排玩家和备用
        ws.cell(row=row_cursor, column=1, value=f"{slot} 已安排玩家角色及其备用角色").font = ws.cell(row=row_cursor,column=1).font.copy(bold=True, color="0000FF")
        row_cursor += 1
        has_assigned_backup = False

        for pname in assigned_pnames:
            assigned_chars = [(cname, cls) for pname2, cname, cls in roster if pname2 == pname]
            remaining_chars = cd_role_pool.get(pname, [])
            available_backup_chars = [char for char in remaining_chars if slot in player_lookup[pname].available_times]

            has_assigned_backup = True
            pname_display = f"{pname}（{player_time_disp[pname]}）"
            ws.cell(row=row_cursor, column=1, value=pname_display)
            col_cursor = 2

            for cname, cls in assigned_chars:
                main_cls = get_classes(cls)[0]
                ws.cell(row=row_cursor, column=col_cursor, value=f"已安排: {cname}-{cls}").fill = PatternFill("solid", fgColor=CLASS_COLORS.get(main_cls,'FFFFFF'))
                col_cursor += 1
            for char in available_backup_chars:
                main_cls = get_classes(char['class'])[0]
                ws.cell(row=row_cursor, column=col_cursor, value=f"备用: {char['name']}-{char['class']}").fill = PatternFill("solid", fgColor=CLASS_COLORS.get(main_cls,'FFFFFF'))
                col_cursor += 1
            row_cursor += 1
        if not has_assigned_backup:
            ws.cell(row=row_cursor, column=1, value="✅无")
            row_cursor += 1
        row_cursor += 2

    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 25
    wb.save(filename)
    print(f"✅团本排表导出成功: {filename}")
    return result

# 定义按ROLES顺序排序的辅助函数
def sort_roles(roster):
    tanks = [r for r in roster if ROLES.get(r[2]) == '坦克']
    melee = [r for r in roster if ROLES.get(r[2]) == '近战输出']
    ranged = [r for r in roster if ROLES.get(r[2]) == '远程输出']
    heals = [r for r in roster if ROLES.get(r[2]) == '治疗']
    return tanks + melee + ranged + heals


def main(input_file, output_file):
    players = read_excel(input_file)
    cd_role_pool = build_cd_role_pool(players)
    time_slots = ["周四-19:30", "周四-20:30", "周五-19:30", "周五-20:30",
                  "周六-19:30", "周六-20:30", "周日-19:30", "周日-20:30",
                  "周一-19:30", "周一-20:30"]

    result = {}
    for slot in time_slots:
        print(f"\n排团时间: {slot}")
        roster = create_group(players, cd_role_pool, slot, time_slots,
                              ["大哥", "薄荷", "古子哥", "戴森", "大力", "可美", "老四", "菲兹", "炸酱面", "白胖", "甜思思", "张三", "狗哥", "故事", "老K", "小K"])
        result[slot] = roster

        # 按坦克->近战->远程->治疗 已经定义的顺序排序
        roster_sorted = sort_roles(roster)

        # 5*5 矩阵输出, 方便终端查看
        print(f"【{slot}】团本名单(5x5展示)：")
        for i in range(5):
            row = roster_sorted[i * 5:(i + 1) * 5]
            row_str = ""
            for pname, char_name, cls in row:
                display_name = f"{pname}-{char_name}-{cls}" if pname else f"空缺-{cls}"
                row_str += f"{display_name:<20}"  # 每个位置占据20字符宽度，整齐美观
            print(row_str)

    # suffix = datetime.now().strftime("%Y%m%d%H%M%S")
    # out_path = '/Users/bytedance/Downloads/团本排表结果' + suffix + '.xlsx'
    return export_schedule_to_excel(result, players, cd_role_pool, output_file)


if __name__ == "__main__":
    suffix = datetime.now().strftime("%Y%m%d%H%M%S")

    input_file = '/Users/bytedance/Downloads/轻风成员.xlsx'
    output_file = '/Users/bytedance/Downloads/团本排表结果' + suffix + '.xlsx'
    main(input_file, output_file)
